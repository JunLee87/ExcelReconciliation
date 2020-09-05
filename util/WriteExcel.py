#!/usr/bin/env python
# coding=UTF-8
"""
@Author: Jun Lee
@LastEditors: Jun Lee
@contact: lijun87@foxmail.com
@file: WriteExcel.py
@LastEditTime: 2020-08-26 15:30
@Descripttion:
对EXCEL写入值
"""
import openpyxl
from openpyxl.styles import PatternFill


class WriteExcel:
    def __init__(self,filePath,sheetName):
        # 读取EXCLE,返回一个工作薄对象
        self.wBook = openpyxl.load_workbook(filePath)
        # 切换到 名字为 的sheet
        self.wSheet = self.wBook[sheetName]

    # 行、列号.行列号都是从1开始
    def setValue(self,row,col,value):
        # 行、列号
        self.wSheet.cell(row, col).value = value

    # 保存文件
    def saveExcel(self,filePath):
        self.wBook.save(filePath)

    # 设置单元格颜色
    def setColour(self,row, col,colour='E39191'):
        fillObj = PatternFill("solid", colour)
        # 指定 某个单元格背景色
        self.wSheet.cell(row, col).fill =fillObj


if __name__ == '__main__':
    filePath='../Resources/测试.xlsx'
    wExcel=WriteExcel(filePath,'sheet1')
    wExcel.setValue(8,21,555888)
    wExcel.setColour(8,21)
    wExcel.saveExcel(filePath)
